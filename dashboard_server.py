import json
import os
import socket
from datetime import datetime, date
from http.server import ThreadingHTTPServer, SimpleHTTPRequestHandler

import openpyxl

PORT = 8766

BASE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(BASE, "Daily Production Report_Aug_26.xlsx")
HTML = os.path.join(BASE, "Cospower_Production_Dashboard_Pylon_v3.html")


def clean(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")

    if isinstance(v, float):
        if v != v or v in (float("inf"), float("-inf")):
            return None
        return v

    if isinstance(v, (int, str, bool)) or v is None:
        return v

    return str(v)


def num(v):
    if isinstance(v, bool) or v is None or v == "":
        return None

    if isinstance(v, (int, float)):
        return float(v)

    try:
        return float(str(v).replace(",", "").strip())
    except Exception:
        return None


def read_data():
    wb = openpyxl.load_workbook(
        EXCEL,
        data_only=True,
        read_only=True
    )

    if (
        "DR + Reactor Status" not in wb.sheetnames
        or "CAP + STRUCTURE Status" not in wb.sheetnames
    ):
        raise RuntimeError("Required Excel sheets are missing.")

    drs = wb["DR + Reactor Status"]
    caps = wb["CAP + STRUCTURE Status"]

    dr = []
    cap = []
    struct = []

    # DR + Reactor
    for row in drs.iter_rows(min_row=5, values_only=True):

        if isinstance(row[0], (datetime, date)):
            dr.append({
                "date": clean(row[0]),
                "project": clean(row[1]),
                "dr": num(row[2]),
                "cubicle": num(row[4]),
                "assembly": num(row[5]),
                "readyTest": num(row[6]),
                "tested": num(row[7]),
                "readyDispatch": num(row[8]),
                "dispatched": num(row[9]),
                "remark": clean(row[10]),
                "reactorWip": num(row[13]),
                "reactorTest": num(row[14]),
                "reactorReady": num(row[15]),
                "reactorDispatched": num(row[16]),
                "reactorRemark": clean(row[17]),
            })

    # CAP + STRUCTURE
    for row in caps.iter_rows(min_row=5, values_only=True):

        if isinstance(row[0], (datetime, date)):
            cap.append({
                "date": clean(row[0]),
                "elements": num(row[1]),
                "encasing": num(row[3]),
                "welding": num(row[5]),
                "vdp": num(row[10]),
                "readyInspection": num(row[12]),
                "readyDispatch": num(row[13]),
                "dispatched": num(row[14]),
                "balance": num(row[15]),
                "remark": clean(row[16]),
            })

        if isinstance(row[18], (datetime, date)):
            struct.append({
                "date": clean(row[18]),
                "black": num(row[19]),
                "toGalv": num(row[20]),
                "galvReady": num(row[21]),
                "dispatched": num(row[22]),
                "remark": clean(row[23]),
            })

    targets = {
        "dr": num(drs.cell(3, 3).value) or 0,
        "cubicle": num(drs.cell(3, 5).value) or 0,
        "assembly": num(drs.cell(3, 6).value) or 0,
        "elements": num(caps.cell(3, 2).value) or 0,
        "encasing": num(caps.cell(3, 4).value) or 0,
        "welding": num(caps.cell(3, 6).value) or 0,
    }

    bf = {
        "drBlock": num(drs.cell(4, 3).value) or 0,
        "drCubicle": num(drs.cell(4, 5).value) or 0,
        "drAssembly": num(drs.cell(4, 6).value) or 0,
        "drReady": num(drs.cell(4, 9).value) or 0,
        "capReady": num(caps.cell(4, 14).value) or 0,
        "capBalance": num(caps.cell(4, 16).value) or 0,
        "structGalvReady": num(caps.cell(4, 22).value) or 0,
    }

    mtime = os.path.getmtime(EXCEL)

    wb.close()

    return {
        "generatedAt": datetime.fromtimestamp(mtime).strftime(
            "%Y-%m-%d %H:%M:%S"
        ),
        "dr": dr,
        "cap": cap,
        "struct": struct,
        "targets": targets,
        "bf": bf,
    }


def send_json(handler, obj, status=200):
    payload = json.dumps(
        obj,
        ensure_ascii=False,
        allow_nan=False,
        separators=(",", ":")
    ).encode("utf-8")

    handler.send_response(status)
    handler.send_header(
        "Content-Type",
        "application/json; charset=utf-8"
    )
    handler.send_header("Cache-Control", "no-store")
    handler.send_header("Content-Length", str(len(payload)))
    handler.end_headers()
    handler.wfile.write(payload)


class Handler(SimpleHTTPRequestHandler):

    def do_GET(self):

        try:

            if self.path.startswith("/api/data"):
                send_json(self, read_data())
                return

            if self.path in (
                "/",
                "/Cospower_Production_Dashboard_Pylon_v3.html",
            ):

                data = read_data()

                with open(HTML, encoding="utf-8") as f:
                    template = f.read()

                marker = (
                    "window.COSPOWER_DATA || "
                    "{dr:[],cap:[],struct:[],targets:{},bf:{}}"
                )

                injected = template.replace(
                    marker,
                    json.dumps(
                        data,
                        ensure_ascii=False,
                        allow_nan=False
                    )
                )

                payload = injected.encode("utf-8")

                self.send_response(200)
                self.send_header(
                    "Content-Type",
                    "text/html; charset=utf-8"
                )
                self.send_header("Cache-Control", "no-store")
                self.send_header(
                    "Content-Length",
                    str(len(payload))
                )
                self.end_headers()
                self.wfile.write(payload)
                return

        except Exception as e:
            send_json(self, {"error": str(e)}, 500)
            return

        return super().do_GET()


if __name__ == "__main__":

    os.chdir(BASE)

    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(("8.8.8.8", 80))
        local_ip = s.getsockname()[0]
        s.close()

    except Exception:
        local_ip = "localhost"

    print("=" * 60)
    print("COSPOWER DASHBOARD IS LIVE!")
    print(f"👉 Link for management: http://{local_ip}:{PORT}/")
    print(f"👉 Link for your PC:    http://127.0.0.1:{PORT}/")
    print(f"Excel file source:     {EXCEL}")
    print("=" * 60)
    print(
        "CRITICAL: Keep this terminal window open while management is using it."
    )

    ThreadingHTTPServer(("", PORT), Handler).serve_forever()